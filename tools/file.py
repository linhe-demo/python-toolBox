# 文件读取类
import json
import os

import pandas as pd
import os.path
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
from PIL import Image as PILImage
import uuid


class File:
    def __init__(self, path=None, row=None, fileData=None, sheetName=None, sheetTitle=None, txtData=None, index=None, target=None):
        self.path = path
        self.fileData = fileData
        self.sheetName = sheetName
        self.sheetTitle = sheetTitle
        self.row = row
        self.txtData = txtData
        self.index = index
        self.target = target

    def read_json_config(self):
        with open(os.path.dirname(os.path.dirname(os.path.realpath(__file__))) + self.path, 'r') as json_file:
            return json.load(json_file)

    def read_excl(self):
        raw_data = pd.read_excel(self.path, header=0)  # header=0表示第一行是表头，就自动去除了
        return raw_data.values

    def read_txt(self):
        data = []
        with open(self.path, 'r', encoding='utf-8') as f:
            for line in f:
                data_line = line.strip("\n").split("\r")  # 去除首尾换行符，并按换行划分
                data.append(data_line)
        return data

    def writeTxt(self):
        if not os.path.exists(self.path):
            open(self.path, 'w', encoding='utf-8').close()

        with open(self.path, 'w', encoding='utf-8') as f:
            for i in self.txtData:
                f.write(i + "\n")

    def writeTxtInline(self):
        if not os.path.exists(self.path):
            open(self.path, 'w', encoding='utf-8').close()

        with open(self.path, 'w', encoding='utf-8') as f:
            f.write(self.txtData + "\n")

    def readExcelSheet(self):
        return pd.read_excel(self.path, sheet_name=self.sheetName).to_dict("records")

    def readRow(self):
        feature = []
        lists = []
        data = self.read_excl()  # 文件位置
        if self.row == 1:
            feature = data[:, 0:1]  # 取第一列
        elif self.row == 2:
            feature = data[:, 1:2]  # 取第二列

        m = 0
        for i in feature:
            tmpKey = str(feature[m][0])
            lists.append(tmpKey)
            m += 1
        return lists

    def writeExcel(self):
        # 检查文件夹是否存在，不存在直接创建
        path = "/".join(self.path.split("/")[:-1])
        if not os.path.exists(path):
            os.makedirs(path)

        # 使用 pandas 将数据写入 Excel
        with pd.ExcelWriter(self.path, engine='openpyxl') as writer:
            for num in self.fileData:
                if len(self.fileData[num]) == 0:
                    continue
                data = pd.DataFrame(self.fileData[num])
                data = data[self.sheetTitle[num]]
                data.to_excel(writer, sheet_name=self.sheetName[num], index=False, header=True)

        # 使用 openpyxl 打开 Excel 并设置行高行宽
        book = load_workbook(self.path)
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            for row in sheet.iter_rows():
                sheet.row_dimensions[row[0].row].height = 100
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width  # 设置列宽

        # 保存并关闭工作簿
        book.save(self.path)
        book.close()

    def urlToImg(self):
        workbook = load_workbook(self.path)
        sheet = workbook[self.sheetName]
        num = 2
        temp_image_paths = []  # 存储临时图片文件路径
        print("开始处理 sheet {} 数据".format(sheet))
        for row in sheet.iter_rows(min_row=2, values_only=True):
            image_link = sheet[f'%s{num}' % self.index].value  # 图片链接
            write_cell = sheet[f'%s{num}' % self.target]  # 插入行
            num += 1
            if image_link:
                response = requests.get(image_link)
                try:
                    image = PILImage.open(BytesIO(response.content))
                except:
                    continue
                image.thumbnail((100, 100))  # 调整图片大小
                # 生成唯一的临时文件名
                temp_image_path = f'../../tmp/temp_image_{uuid.uuid4().hex}.jpg'

                image.save(temp_image_path, 'JPEG')

                # 将图片插入Excel单元格
                img = Image(temp_image_path)
                sheet.add_image(img, write_cell.coordinate)

                # 设置行高为图片高度
                row_height = image.size[1]
                sheet.row_dimensions[write_cell.row].height = row_height

                temp_image_paths.append(temp_image_path)  # 添加临时图片文件路径

        workbook.save(self.path)

        # 删除所有临时图片文件
        for temp_image_path in temp_image_paths:
            if os.path.exists(temp_image_path):
                os.remove(temp_image_path)

